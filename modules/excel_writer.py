"""
excel_writer.py
Matches the original Energybae template EXACTLY.

One "pair block" = 2 consumers side by side:
  B  = Sr.No label col   (left)
  C  = Month             (left)
  D  = Units             (left)
  E  = Bill Amount       (left)
  F  = Unit Cost         (left)
  G  = gap
  H  = Month             (right)
  I  = Units             (right)
  J  = Bill Amount       (right)
  K  = Unit Cost         (right)

Info rows (1-5):
  B  = Label (orange)
  C:E merged = left consumer value
  G (blank gap)
  H  = Label (orange)
  I:K merged = right consumer value
"""

import os, math
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ORANGE    = "C55A11"
YELLOW    = "FFFF00"
AMBER     = "FFC000"
GREEN     = "00B050"
RED       = "FF0000"
LIGHT_GRN = "E2EFDA"
WHITE     = "FFFFFF"
BLACK     = "000000"

MASTER    = "Solar_Report_Master.xlsx"
# Each pair block = 11 cols (B..L) + 1 gap = 12 cols
BLOCK_W   = 12
START_COL = 2   # col B


def _fill(c): return PatternFill(start_color=c, end_color=c, fill_type="solid")
def _bdr():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _calc(bill):
    uv = []
    for e in bill.get("monthly_history", []):
        u = e.get("units", "")
        try:
            if u not in ("", None): uv.append(float(u))
        except: pass
    avg = round(sum(uv)/len(uv), 2) if uv else float(bill.get("units_consumed", 0) or 0)
    kw  = avg / (30 * 5 * 0.8)
    raw = kw / 0.4
    rec = math.ceil(kw * 2) / 2
    npn = math.ceil(rec / 0.4)
    return avg, kw, raw, rec, npn

def _consumer_key(bill):
    n = str(bill.get("consumer_number", "") or "").strip()
    return n if n else str(bill.get("consumer_name", "") or "").strip().lower()

def _pair_key(b1, b2):
    return "||".join(sorted([_consumer_key(b1), _consumer_key(b2)]))

def _unmerge_cell(ws, row, col):
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            try: ws.unmerge_cells(str(mr))
            except: pass

def _S(ws, row, col, val="", bold=False, color=BLACK, bg=None, align="center", nfmt=None):
    _unmerge_cell(ws, row, col)
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(bold=bold, size=10, color=color)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if bg: c.fill = _fill(bg)
    c.border = _bdr()
    if nfmt: c.number_format = nfmt
    return c

def _M(ws, row, c1, c2, val="", bold=False, color=BLACK, bg=None, align="center"):
    for col in range(c1, c2+1):
        _unmerge_cell(ws, row, col)
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=val)
    c.font      = Font(bold=bold, size=10, color=color)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    if bg: c.fill = _fill(bg)
    c.border = _bdr()
    return c


def _write_pair(ws, bill1, bill2, sc):
    """
    Write a consumer pair block starting at column sc.
    sc+0  = Sr.No / label
    sc+1  = Month L
    sc+2  = Units L
    sc+3  = Bill Amount L
    sc+4  = Unit Cost L
    sc+5  = gap
    sc+6  = Month R
    sc+7  = Units R
    sc+8  = Bill Amount R
    sc+9  = Unit Cost R
    sc+10 = (spare / gap)
    sc+11 = gap col
    """
    S = lambda r, c, v="", **kw: _S(ws, r, c, v, **kw)
    M = lambda r, c1, c2, v="", **kw: _M(ws, r, c1, c2, v, **kw)

    # Column widths
    col_w = [6, 16, 10, 13, 10, 3, 16, 10, 13, 10, 3, 3]
    for i, w in enumerate(col_w):
        ws.column_dimensions[get_column_letter(sc+i)].width = w

    # Row heights
    ws.row_dimensions[8].height = 28

    cons_no1 = str(bill1.get("consumer_number", "") or "").strip()
    cons_no2 = str(bill2.get("consumer_number", "") or "").strip()
    load1    = bill1.get("sanctioned_load_kw", "")
    load2    = bill2.get("sanctioned_load_kw", "")

    # ── ROWS 1-5: consumer info ───────────────────────────────────────────────
    labels = ["Consumer Name", "Consumer No", "Fixed Charges",
              "Sanct. Load (kW)", "Connection Type"]
    vals1  = [
        bill1.get("consumer_name", ""),
        cons_no1,
        bill1.get("fixed_charges", 130) or 130,
        f"{load1}KW" if load1 else "",
        bill1.get("tariff_category", "90/LT I Res 1-Phase"),
    ]
    vals2  = [
        bill2.get("consumer_name", ""),
        cons_no2,
        bill2.get("fixed_charges", 130) or 130,
        f"{load2}KW" if load2 else "",
        bill2.get("tariff_category", "90/LT I Res 1-Phase"),
    ]

    for i, (lbl, v1, v2) in enumerate(zip(labels, vals1, vals2), 1):
        S(i, sc,   lbl, bold=True, bg=ORANGE, color=WHITE)
        M(i, sc+1, sc+4, v1)
        S(i, sc+5, "")                          # gap
        S(i, sc+6, lbl, bold=True, bg=ORANGE, color=WHITE)
        M(i, sc+7, sc+10, v2)

    # Force consumer numbers as text
    for row, col, cno in [(2, sc+1, cons_no1), (2, sc+7, cons_no2)]:
        c = ws.cell(row=row, column=col, value=cno)
        c.number_format = "@"
        c.alignment = Alignment(horizontal="center", vertical="center")

    # ── ROW 6: contract demand ────────────────────────────────────────────────
    M(6, sc, sc+4, "Contract Demand (KVA) :", bold=True, align="left")
    S(6, sc+5, "")

    # ── ROW 7: solar panel uses ───────────────────────────────────────────────
    S(7, sc,   "Solar Pannel uses", bold=True, bg=YELLOW, color=BLACK)
    S(7, sc+1, 600, bold=True, bg=YELLOW, color=BLACK)
    for col in [sc+2, sc+3, sc+4, sc+5]: S(7, col, "")

    # ── ROW 8: table headers ──────────────────────────────────────────────────
    S(8, sc,    "Sr.No",       bold=True, bg=ORANGE, color=WHITE)
    S(8, sc+1,  "Month",       bold=True, bg=ORANGE, color=WHITE)
    S(8, sc+2,  "Units",       bold=True, bg=ORANGE, color=WHITE)
    S(8, sc+3,  "Bill Amount", bold=True, bg=ORANGE, color=WHITE)
    S(8, sc+4,  "Unit\nCost",  bold=True, bg=ORANGE, color=WHITE)
    S(8, sc+5,  "",            bold=True, bg=ORANGE, color=WHITE)  # gap header
    S(8, sc+6,  "Month",       bold=True, bg=ORANGE, color=WHITE)
    S(8, sc+7,  "Units",       bold=True, bg=ORANGE, color=WHITE)
    S(8, sc+8,  "Bill Amount", bold=True, bg=ORANGE, color=WHITE)
    S(8, sc+9,  "Unit\nCost",  bold=True, bg=ORANGE, color=WHITE)
    S(8, sc+10, "",            bold=True, bg=ORANGE, color=WHITE)

    # ── ROWS 9-21: monthly history ────────────────────────────────────────────
    h1 = list(bill1.get("monthly_history", []))
    h2 = list(bill2.get("monthly_history", []))
    while len(h1) < 13: h1.append({"month":"","units":"","bill_amount":"","unit_cost":""})
    while len(h2) < 13: h2.append({"month":"","units":"","bill_amount":"","unit_cost":""})

    def _dv(v):  # display value — blank if 0
        try: return v if float(v) != 0 else ""
        except: return "" if v in (None, "") else v

    for i in range(13):
        row = 9 + i
        e1, e2 = h1[i], h2[i]
        u1 = e1.get("units","");      u2 = e2.get("units","")
        b1 = _dv(e1.get("bill_amount","")); b2 = _dv(e2.get("bill_amount",""))
        c1 = _dv(e1.get("unit_cost","")); c2 = _dv(e2.get("unit_cost",""))

        S(row, sc,    i+2,                                    align="center")
        S(row, sc+1,  e1.get("month",""),                     align="center")
        S(row, sc+2,  u1 if u1 not in ("",None) else "",      align="center")
        S(row, sc+3,  b1,                                     align="center")
        S(row, sc+4,  c1,                                     align="center")
        S(row, sc+5,  "",                                     align="center")
        S(row, sc+6,  e2.get("month",""),                     align="center")
        S(row, sc+7,  u2 if u2 not in ("",None) else "",      align="center")
        S(row, sc+8,  b2,                                     align="center")
        S(row, sc+9,  c2,                                     align="center")
        S(row, sc+10, "",                                     align="center")

    # ── Calculations ──────────────────────────────────────────────────────────
    avg1,kw1,sp1,rec1,np1 = _calc(bill1)
    avg2,kw2,sp2,rec2,np2 = _calc(bill2)

    def stat_block(avg, kw, sp, rec, npn, col_lbl, col_val):
        S(22, col_lbl, "Average",        bold=True, align="left")
        S(22, col_val, avg,              align="center")
        S(23, col_lbl, "kW",             bold=True, align="left")
        S(23, col_val, round(kw,9),      align="center")
        S(24, col_lbl, "Solar Panels",   bold=True, align="left")
        S(24, col_val, round(sp,9),      align="center")
        S(25, col_lbl, "Solar capacity", bold=True, bg=AMBER, align="left")
        S(25, col_val, rec, bold=True,   bg=GREEN, color=WHITE, align="center")
        S(26, col_lbl, "Number of\nPanels", bold=True, bg=AMBER, align="left")
        S(26, col_val, npn, bold=True,   bg=RED, color=WHITE, align="center")

    stat_block(avg1,kw1,sp1,rec1,np1, sc+1,   sc+2)
    stat_block(avg2,kw2,sp2,rec2,np2, sc+6,   sc+7)

    # Fill blank stat cells
    for row in range(22, 27):
        for col in [sc, sc+3, sc+4, sc+5, sc+8, sc+9, sc+10]:
            S(row, col, "")

    # ── ROWS 29-30: totals ────────────────────────────────────────────────────
    for row, lbl, val in [
        (29, "Total solar capacity",   round(rec1+rec2, 1)),
        (30, "Number of solar panels", np1+np2),
    ]:
        c = ws.cell(row=row, column=sc+1, value=lbl)
        c.font = Font(bold=True, size=10)
        c.alignment = Alignment(horizontal="left")
        t = ws.cell(row=row, column=sc+2, value=val)
        t.font = Font(bold=True, size=10)
        t.fill = _fill(LIGHT_GRN)
        t.alignment = Alignment(horizontal="center")
        t.border = _bdr()

    # ── Pair key at row 32 (white text — invisible but readable by openpyxl) ──
    key_val = _pair_key(bill1, bill2)
    kc = ws.cell(row=32, column=sc, value=key_val)
    kc.font = Font(color="FFFFFF", size=8)
    # Also store in row 33 as plain backup
    kc2 = ws.cell(row=33, column=sc, value=key_val)
    kc2.font = Font(color="FFFFFF", size=8)


# ── PUBLIC: individual report ─────────────────────────────────────────────────

def build_excel(bill, solar, output_dir="outputs", bill2=None, solar2=None):
    os.makedirs(output_dir, exist_ok=True)
    b2 = bill2 if bill2 else bill   # mirror same consumer if only 1 uploaded
    wb = Workbook()
    ws = wb.active
    ws.title = "Solar Load Calculator"
    _write_pair(ws, bill, b2, sc=START_COL)
    fname = f"Solar_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    fpath = os.path.join(output_dir, fname)
    wb.save(fpath)
    return fpath, fname


# ── PUBLIC: master report ─────────────────────────────────────────────────────

def build_combined_excel(bill, solar, output_dir="outputs", bill2=None, solar2=None):
    os.makedirs(output_dir, exist_ok=True)
    b2 = bill2 if bill2 else bill
    master_path = os.path.join(output_dir, MASTER)
    new_key     = _pair_key(bill, b2)

    if os.path.exists(master_path):
        wb = load_workbook(master_path)
        ws = wb.active

        # Collect ALL existing keys first
        existing_keys = set()
        idx = 0
        last_used_idx = -1
        while True:
            sc  = START_COL + idx * BLOCK_W
            # Read key from row 32 or 33 (stored in both as backup)
            key32 = ws.cell(row=32, column=sc).value
            key33 = ws.cell(row=33, column=sc).value
            key_cell = key32 or key33
            # Check if this block has any data
            name_cell = ws.cell(row=1, column=sc+1).value
            if name_cell is None and key_cell is None:
                break   # empty slot found
            if key_cell:
                existing_keys.add(str(key_cell).strip())
            last_used_idx = idx
            idx += 1

        next_idx = last_used_idx + 1

        # Check if this pair already exists
        if new_key in existing_keys:
            return master_path, MASTER   # duplicate → skip

        _write_pair(ws, bill, b2, sc=START_COL + next_idx * BLOCK_W)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Solar Load Calculator"
        _write_pair(ws, bill, b2, sc=START_COL)

    wb.save(master_path)
    return master_path, MASTER